from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
import time
from selenium.webdriver.chrome.options import Options

# Abre o navegador em modo anônimo e entra no site de consulta
def start_driver():
    options = Options()
    options.add_argument('--incognito')  # abre no modo anônimo
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.get("https://www.consultacnpj.com/")
    return driver

# Carrega a planilha com os CNPJs
wb = openpyxl.load_workbook('C:\\Users\\guilherme.visconi\\Desktop\\CNPJOTO\\CNPJ.xlsx')  

# Cria uma planilha nova para salvar os resultados
results_wb = openpyxl.Workbook()
results_sheet = results_wb.active
results_sheet.append(['CNPJ', 'Porte', 'Nome Empresarial'])

# Passa por todos os CNPJs da planilha
sheet = wb.active
for row in sheet.iter_rows(min_row=2, max_col=1):  # começa na linha 2 (linha 1 é cabeçalho)
    cnpj = row[0].value
    if not cnpj:  # se não tiver CNPJ, pula pra próxima
        continue

    print(f"Iniciando consulta para CNPJ: {cnpj}")

    driver = start_driver()

    try:
        # Clica no botão de aceitar os termos
        botao_aceitar_termos = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), 'Aceito os Termos')]"))
        )
        botao_aceitar_termos.click()
        print("Aceitei os termos.")

        # Espera o campo de CNPJ aparecer e preenche
        campo_cnpj = WebDriverWait(driver, 5).until(
            EC.visibility_of_element_located((By.XPATH, "//input[@placeholder='00.000.000/0000-00']"))
        )
        campo_cnpj.send_keys(cnpj)
        print("CNPJ preenchido.")

        # Dá um tempo para o site processar
        time.sleep(3)

        # Espera a página carregar os dados
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//p[contains(text(), 'Porte estimado')]"))
        )

        # Busca os blocos de informações
        divs_porte = driver.find_elements(By.XPATH, "//div[@class='crawler-dashboard-item p4']")
        
        if len(divs_porte) >= 7:
            porte_texto = divs_porte[6].text  # o porte fica no 7º bloco
            print(f"Porte Estimado: {porte_texto}")
            
            # Tipos de empresa possíveis
            tipos_de_empresa = ["DEMAIS", "ME", "MEI", "EPP", "LTDA", "S/A", "SOCIEDADE", "EIRELI"]
            
            tipo_encontrado = None
            for tipo in tipos_de_empresa:
                if tipo in porte_texto:
                    tipo_encontrado = tipo
                    break

            if tipo_encontrado:
                print(f"Tipo de empresa: {tipo_encontrado}")
            else:
                print("Não consegui identificar o tipo de empresa.")
        else:
            print("Não encontrei o bloco do porte.")

        # Pega o Nome Empresarial
        nome_empresarial_elemento = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//div[@id='company-data']//div[@class='p4 print-border']//label[contains(text(), 'Nome empresarial:')]//following-sibling::p[@class='font-color-tertiary']"))
        )
        nome_empresarial = nome_empresarial_elemento.text.strip()
        print(f"Nome Empresarial: {nome_empresarial}")

        # Salva os resultados
        results_sheet.append([cnpj, porte_texto, nome_empresarial])
        print(f"Resultado salvo para {cnpj}.")

    except Exception as e:
        print(f"Erro durante a consulta: {e}")
    finally:
        # Fecha o navegador depois de esperar um pouco
        print("Consulta finalizada. Aguardando antes da próxima...")
        time.sleep(60)  
        driver.quit()  

# Salva a planilha com os resultados
results_wb.save('resultados_consulta.xlsx')
